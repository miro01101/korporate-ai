from pathlib import Path
import sys

from openpyxl import load_workbook


if len(sys.argv) != 2:
    raise SystemExit("Pouzitie: show-data-contract.py /cesta/subor.xlsx")

path = Path(sys.argv[1])

workbook = load_workbook(
    filename=path,
    read_only=True,
    data_only=False,
)

worksheet = workbook["data_dictionary"]

print("sheet_name | column_name | data_type | required | description")
print("-" * 120)

for row in worksheet.iter_rows(min_row=2, values_only=True):
    values = ["" if value is None else str(value) for value in row]
    print(" | ".join(values))

workbook.close()

print()
print("DATA_CONTRACT_READ_OK=ANO")
