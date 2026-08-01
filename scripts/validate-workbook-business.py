from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import sys

from openpyxl import load_workbook
from workbook_normalization import normalize_workbook

MAX_EXAMPLES = 20
HEADERS = {
    "products": "product_id product_name category unit purchase_price sales_price supplier minimum_order_quantity lead_time_days weight_kg volume_m3".split(),
    "sales": "order_id order_date product_id quantity unit_price customer_id customer_name region order_status expedition_date".split(),
    "inventory": "snapshot_date product_id stock_actual stock_reserved stock_available warehouse_location min_stock max_stock".split(),
    "purchases": "purchase_order_id order_date delivery_date supplier product_id ordered_quantity delivered_quantity purchase_price".split(),
    "expedition": "order_id received_at picked_at expedition_date delivery_type vehicle_id region weight_kg volume_m3".split(),
    "vehicles": "vehicle_id capacity_kg capacity_m3 availability cost_per_km driver".split(),
}
REGIONS = {"BA", "BB", "KE", "NR", "PO", "TN", "TT", "ZA"}
STATUSES = {"vybavená"}
DELIVERY_TYPES = {"vlastná", "externá", "osobný odber"}
AVAILABILITY = {"pondelok-piatok", "pondelok-sobota", "podľa objednávky"}
UNITS = {"ks", "bal", "rolka", "vedro", "sada", "m2", "m"}

if len(sys.argv) != 2:
    print("Pouzitie: validate-workbook-business.py /cesta/subor.xlsx", file=sys.stderr)
    raise SystemExit(2)
path = Path(sys.argv[1])
if not path.is_file():
    print(f"Subor neexistuje: {path}", file=sys.stderr)
    raise SystemExit(2)

counts = Counter()
examples = defaultdict(list)


def issue(level, code, sheet, row, key, field, value, expected):
    marker = (level, code)
    counts[marker] += 1
    if len(examples[marker]) < MAX_EXAMPLES:
        location = sheet if row is None else f"{sheet}!{row}"
        examples[marker].append(
            f"{location}: key={key!r}; field={field}; value={value!r}; expected={expected}"
        )


def err(code, sheet, row, key, field, value, expected):
    issue("ERROR", code, sheet, row, key, field, value, expected)


def warn(code, sheet, row, key, field, value, expected):
    issue("WARNING", code, sheet, row, key, field, value, expected)


def num(value):
    if value is None or isinstance(value, bool):
        return None
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    return result if result.is_finite() else None


def day(value):
    if isinstance(value, datetime):
        return value.date()
    return value if isinstance(value, date) else None


def rows_from(workbook, sheet):
    worksheet = workbook[sheet]
    actual = [
        "" if value is None else str(value)
        for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    if actual != HEADERS[sheet]:
        print(
            f"Neplatna hlavicka {sheet}: ocakavane={HEADERS[sheet]!r}; skutocne={actual!r}",
            file=sys.stderr,
        )
        raise SystemExit(2)
    result = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        row = dict(zip(actual, values, strict=True))
        row["_row"] = row_number
        result.append(row)
    return result


def duplicate_keys(rows, sheet, fields, code):
    seen = {}
    for row in rows:
        key = tuple(row[field] for field in fields)
        if key in seen:
            err(code, sheet, row["_row"], key, ",".join(fields), key,
                f"unikatna kombinacia; prvy riadok {seen[key]}")
        else:
            seen[key] = row["_row"]


def exact_duplicates(rows, sheet, code):
    seen = {}
    fields = HEADERS[sheet]
    for row in rows:
        signature = tuple(row[field] for field in fields)
        if signature in seen:
            err(code, sheet, row["_row"], signature[0], "row", "uplna duplicita",
                f"unikatny riadok; prvy riadok {seen[signature]}")
        else:
            seen[signature] = row["_row"]


def bounds(row, sheet, key, positive=(), nonnegative=()):
    for field in positive:
        value = num(row[field])
        if value is not None and value <= 0:
            err("BIZ-E030", sheet, row["_row"], key, field, row[field], "hodnota > 0")
    for field in nonnegative:
        value = num(row[field])
        if value is not None and value < 0:
            err("BIZ-E031", sheet, row["_row"], key, field, row[field], "hodnota >= 0")


def header_consistency(grouped, sheet, fields, code):
    for key, group in grouped.items():
        reference = group[0]
        for row in group[1:]:
            for field in fields:
                if row[field] != reference[field]:
                    err(code, sheet, row["_row"], key, field, row[field],
                        f"hodnota prveho riadka objednavky {reference[field]!r}")


def repeated_product_warning(grouped, sheet, order_field, code):
    for order_id, group in grouped.items():
        product_counts = Counter(row["product_id"] for row in group)
        for product_id, count in product_counts.items():
            if count > 1:
                warn(code, sheet, None, order_id, "product_id", product_id,
                     "produkt najviac raz v objednavke alebo business zdovodnenie")


try:
    workbook = load_workbook(path, read_only=False, data_only=False)
    normalize_workbook(workbook)
except Exception as exc:
    print(f"Workbook sa nepodarilo otvorit: {exc}", file=sys.stderr)
    raise SystemExit(2) from exc

missing = set(HEADERS) - set(workbook.sheetnames)
if missing:
    workbook.close()
    print(f"Chybaju povinne sheety: {sorted(missing)}", file=sys.stderr)
    raise SystemExit(2)
try:
    data = {sheet: rows_from(workbook, sheet) for sheet in HEADERS}
finally:
    workbook.close()

products = data["products"]
sales = data["sales"]
inventory = data["inventory"]
purchases = data["purchases"]
expedition = data["expedition"]
vehicles = data["vehicles"]

duplicate_keys(products, "products", ("product_id",), "BIZ-E001")
duplicate_keys(vehicles, "vehicles", ("vehicle_id",), "BIZ-E002")
duplicate_keys(inventory, "inventory", ("snapshot_date", "product_id"), "BIZ-E003")
duplicate_keys(expedition, "expedition", ("order_id",), "BIZ-E004")
exact_duplicates(sales, "sales", "BIZ-E005")
exact_duplicates(purchases, "purchases", "BIZ-E006")

product_by_id = {row["product_id"]: row for row in products}
vehicle_by_id = {row["vehicle_id"]: row for row in vehicles}

for row in products:
    key = row["product_id"]
    bounds(row, "products", key, ("minimum_order_quantity",),
           ("purchase_price", "sales_price", "lead_time_days", "weight_kg", "volume_m3"))
    if row["unit"] not in UNITS:
        warn("BIZ-W020", "products", row["_row"], key, "unit", row["unit"], sorted(UNITS))
    if num(row["sales_price"]) is not None and num(row["purchase_price"]) is not None \
            and num(row["sales_price"]) < num(row["purchase_price"]):
        warn("BIZ-W005", "products", row["_row"], key, "sales_price", row["sales_price"],
             "sales_price >= purchase_price")
    zero_warning_codes = {
        "purchase_price": "BIZ-W002",
        "sales_price": "BIZ-W002",
        "weight_kg": "BIZ-W003",
        "volume_m3": "BIZ-W003",
        "lead_time_days": "BIZ-W004",
    }
    for field, code in zero_warning_codes.items():
        if num(row[field]) == 0:
            warn(code, "products", row["_row"], key, field, row[field],
                 "nenulova hodnota alebo zdovodnenie")

for row in vehicles:
    key = row["vehicle_id"]
    bounds(row, "vehicles", key, ("capacity_kg", "capacity_m3"), ("cost_per_km",))
    if row["availability"] not in AVAILABILITY:
        err("BIZ-E063", "vehicles", row["_row"], key, "availability",
            row["availability"], sorted(AVAILABILITY))
    if num(row["cost_per_km"]) == 0:
        warn("BIZ-W004", "vehicles", row["_row"], key, "cost_per_km", 0,
             "nenulova hodnota alebo zdovodnenie")

sales_by_order = defaultdict(list)
names_by_customer = defaultdict(set)
customers_by_name = defaultdict(set)
for row in sales:
    order_id = row["order_id"]
    sales_by_order[order_id].append(row)
    names_by_customer[row["customer_id"]].add(row["customer_name"])
    customers_by_name[row["customer_name"]].add(row["customer_id"])
    if row["product_id"] not in product_by_id:
        err("BIZ-E010", "sales", row["_row"], order_id, "product_id", row["product_id"],
            "existujuci products.product_id")
    bounds(row, "sales", order_id, ("quantity",), ("unit_price",))
    if num(row["unit_price"]) == 0:
        warn("BIZ-W002", "sales", row["_row"], order_id, "unit_price", 0,
             "nenulova hodnota alebo zdovodnenie")
    if day(row["order_date"]) and day(row["expedition_date"]) \
            and day(row["expedition_date"]) < day(row["order_date"]):
        err("BIZ-E040", "sales", row["_row"], order_id, "expedition_date",
            row["expedition_date"], f">= order_date {row['order_date']}")
    if row["region"] not in REGIONS:
        err("BIZ-E060", "sales", row["_row"], order_id, "region", row["region"], sorted(REGIONS))
    if row["order_status"] not in STATUSES:
        err("BIZ-E061", "sales", row["_row"], order_id, "order_status",
            row["order_status"], sorted(STATUSES))

header_consistency(sales_by_order, "sales",
                   ("order_date", "customer_id", "customer_name", "region", "order_status", "expedition_date"),
                   "BIZ-E050")
repeated_product_warning(sales_by_order, "sales", "order_id", "BIZ-W010")
for customer_id, names in names_by_customer.items():
    if len(names) > 1:
        warn("BIZ-W006", "sales", None, customer_id, "customer_name", sorted(map(str, names)),
             "jeden nazov pre customer_id")
for customer_name, ids in customers_by_name.items():
    if len(ids) > 1:
        warn("BIZ-W007", "sales", None, customer_name, "customer_id", sorted(map(str, ids)),
             "jeden customer_id pre nazov")

snapshot_products = defaultdict(set)
snapshot_dates = set()
for row in inventory:
    snapshot = day(row["snapshot_date"])
    key = (snapshot, row["product_id"])
    if row["product_id"] not in product_by_id:
        err("BIZ-E011", "inventory", row["_row"], key, "product_id", row["product_id"],
            "existujuci products.product_id")
    for field in ("stock_actual", "stock_reserved", "stock_available", "min_stock", "max_stock"):
        value = num(row[field])
        if value is not None and value < 0:
            err("BIZ-E023", "inventory", row["_row"], key, field, row[field], "hodnota >= 0")
    actual, reserved, available = map(num, (row["stock_actual"], row["stock_reserved"], row["stock_available"]))
    minimum, maximum = map(num, (row["min_stock"], row["max_stock"]))
    if None not in (actual, reserved, available) and available != actual - reserved:
        err("BIZ-E020", "inventory", row["_row"], key, "stock_available", available,
            f"stock_actual - stock_reserved = {actual - reserved}")
    if actual is not None and reserved is not None and reserved > actual:
        err("BIZ-E021", "inventory", row["_row"], key, "stock_reserved", reserved,
            f"<= stock_actual {actual}")
    if minimum is not None and maximum is not None and minimum > maximum:
        err("BIZ-E022", "inventory", row["_row"], key, "min_stock", minimum,
            f"<= max_stock {maximum}")
    if snapshot:
        snapshot_dates.add(snapshot)
        snapshot_products[snapshot].add(row["product_id"])
        if snapshot.day != 1:
            warn("BIZ-W013", "inventory", row["_row"], key, "snapshot_date", snapshot,
                 "prvy den mesiaca")

all_products = set(product_by_id)
for snapshot, ids in sorted(snapshot_products.items()):
    missing_products = all_products - ids
    if missing_products:
        warn("BIZ-W012", "inventory", None, snapshot, "product_id",
             f"chyba {len(missing_products)} produktov", "vsetky products.product_id")
if snapshot_dates:
    current = (min(snapshot_dates).year, min(snapshot_dates).month)
    last = (max(snapshot_dates).year, max(snapshot_dates).month)
    present = {(value.year, value.month) for value in snapshot_dates}
    while current <= last:
        if current not in present:
            warn("BIZ-W014", "inventory", None, f"{current[0]:04d}-{current[1]:02d}",
                 "snapshot_date", "chyba", "jeden snapshot za mesiac")
        current = (current[0] + 1, 1) if current[1] == 12 else (current[0], current[1] + 1)

purchases_by_order = defaultdict(list)
for row in purchases:
    order_id = row["purchase_order_id"]
    purchases_by_order[order_id].append(row)
    product = product_by_id.get(row["product_id"])
    if product is None:
        err("BIZ-E012", "purchases", row["_row"], order_id, "product_id", row["product_id"],
            "existujuci products.product_id")
    bounds(row, "purchases", order_id, ("ordered_quantity",),
           ("delivered_quantity", "purchase_price"))
    ordered, delivered = num(row["ordered_quantity"]), num(row["delivered_quantity"])
    if ordered is not None and delivered is not None and delivered > ordered:
        err("BIZ-E032", "purchases", row["_row"], order_id, "delivered_quantity", delivered,
            f"<= ordered_quantity {ordered}")
    if day(row["order_date"]) and day(row["delivery_date"]) \
            and day(row["delivery_date"]) < day(row["order_date"]):
        err("BIZ-E041", "purchases", row["_row"], order_id, "delivery_date",
            row["delivery_date"], f">= order_date {row['order_date']}")
    if num(row["purchase_price"]) == 0:
        warn("BIZ-W002", "purchases", row["_row"], order_id, "purchase_price", 0,
             "nenulova hodnota alebo zdovodnenie")
    if product:
        if row["supplier"] != product["supplier"]:
            warn("BIZ-W008", "purchases", row["_row"], order_id, "supplier", row["supplier"],
                 f"primarny dodavatel {product['supplier']!r} alebo schvalena alternativa")
        minimum = num(product["minimum_order_quantity"])
        if ordered is not None and minimum and ordered % minimum:
            warn("BIZ-W009", "purchases", row["_row"], order_id, "ordered_quantity", ordered,
                 f"nasobok minimum_order_quantity {minimum}")

header_consistency(purchases_by_order, "purchases", ("order_date",), "BIZ-E051")
for purchase_order_id, group in purchases_by_order.items():
    suppliers = sorted({str(row["supplier"]) for row in group})
    if len(suppliers) > 1:
        warn(
            "BIZ-W021",
            "purchases",
            None,
            purchase_order_id,
            "supplier",
            suppliers,
            "automaticka normalizacia na source_purchase_order_id + supplier",
        )
repeated_product_warning(purchases_by_order, "purchases", "purchase_order_id", "BIZ-W011")

expedition_by_order = {}
for row in expedition:
    order_id = row["order_id"]
    expedition_by_order.setdefault(order_id, row)
    if order_id not in sales_by_order:
        err("BIZ-E013", "expedition", row["_row"], order_id, "order_id", order_id,
            "existujuci sales.order_id")
    vehicle_id = row["vehicle_id"]
    if vehicle_id is not None and vehicle_id not in vehicle_by_id:
        err("BIZ-E014", "expedition", row["_row"], order_id, "vehicle_id", vehicle_id,
            "existujuci vehicles.vehicle_id")
    bounds(row, "expedition", order_id, (), ("weight_kg", "volume_m3"))
    if row["delivery_type"] not in DELIVERY_TYPES:
        err("BIZ-E062", "expedition", row["_row"], order_id, "delivery_type",
            row["delivery_type"], sorted(DELIVERY_TYPES))
    if row["region"] not in REGIONS:
        err("BIZ-E060", "expedition", row["_row"], order_id, "region", row["region"], sorted(REGIONS))
    if row["delivery_type"] == "vlastná" and vehicle_id is None:
        err("BIZ-E070", "expedition", row["_row"], order_id, "vehicle_id", None,
            "vyplnene vehicle_id")
    if row["delivery_type"] in {"externá", "osobný odber"} and vehicle_id is not None:
        warn("BIZ-W015", "expedition", row["_row"], order_id, "vehicle_id", vehicle_id,
             "prazdna hodnota alebo zdovodnenie")
    received = row["received_at"] if isinstance(row["received_at"], datetime) else None
    picked = row["picked_at"] if isinstance(row["picked_at"], datetime) else None
    shipped = day(row["expedition_date"])
    if received and picked and picked < received:
        err("BIZ-E042", "expedition", row["_row"], order_id, "picked_at", picked,
            f">= received_at {received}")
    if received and shipped and shipped < received.date():
        err("BIZ-E044", "expedition", row["_row"], order_id, "expedition_date", shipped,
            f">= received_at {received.date()}")
    if picked and shipped and shipped < picked.date():
        err("BIZ-E045", "expedition", row["_row"], order_id, "expedition_date", shipped,
            f">= picked_at {picked.date()}")
    vehicle = vehicle_by_id.get(vehicle_id)
    if row["delivery_type"] == "vlastná" and vehicle:
        weight, capacity_kg = num(row["weight_kg"]), num(vehicle["capacity_kg"])
        volume, capacity_m3 = num(row["volume_m3"]), num(vehicle["capacity_m3"])
        if weight is not None and capacity_kg is not None and weight > capacity_kg:
            warn("BIZ-W016", "expedition", row["_row"], order_id, "weight_kg", row["weight_kg"],
                 f"<= capacity_kg {vehicle['capacity_kg']}")
        if volume is not None and capacity_m3 is not None and volume > capacity_m3:
            warn("BIZ-W017", "expedition", row["_row"], order_id, "volume_m3", row["volume_m3"],
                 f"<= capacity_m3 {vehicle['capacity_m3']}")

for order_id, group in sales_by_order.items():
    sale = group[0]
    shipment = expedition_by_order.get(order_id)
    if sale["order_status"] == "vybavená" and shipment is None:
        err("BIZ-E015", "sales", sale["_row"], order_id, "order_id", order_id,
            "prave jeden expedition riadok")
        continue
    if shipment is None:
        continue
    received = shipment["received_at"] if isinstance(shipment["received_at"], datetime) else None
    if received and day(sale["order_date"]) and received.date() < day(sale["order_date"]):
        err("BIZ-E043", "expedition", shipment["_row"], order_id, "received_at", received,
            f">= sales.order_date {sale['order_date']}")
    if day(sale["expedition_date"]) != day(shipment["expedition_date"]):
        err("BIZ-E080", "expedition", shipment["_row"], order_id, "expedition_date",
            shipment["expedition_date"], f"sales.expedition_date {sale['expedition_date']}")
    if sale["region"] != shipment["region"]:
        err("BIZ-E081", "expedition", shipment["_row"], order_id, "region", shipment["region"],
            f"sales.region {sale['region']!r}")
    expected_weight = Decimal(0)
    expected_volume = Decimal(0)
    calculable = True
    for line in group:
        product = product_by_id.get(line["product_id"])
        quantity = num(line["quantity"])
        if not product or quantity is None or num(product["weight_kg"]) is None \
                or num(product["volume_m3"]) is None:
            calculable = False
            break
        expected_weight += quantity * num(product["weight_kg"])
        expected_volume += quantity * num(product["volume_m3"])
    if calculable:
        expected_weight = expected_weight.quantize(Decimal("0.01"))
        expected_volume = expected_volume.quantize(Decimal("0.001"))
        actual_weight = num(shipment["weight_kg"])
        actual_volume = num(shipment["volume_m3"])
        if actual_weight is not None and actual_weight.quantize(Decimal("0.01")) != expected_weight:
            warn("BIZ-W018", "expedition", shipment["_row"], order_id, "weight_kg", actual_weight,
                 f"vypocitana hmotnost {expected_weight}")
        if actual_volume is not None and actual_volume.quantize(Decimal("0.001")) != expected_volume:
            warn("BIZ-W019", "expedition", shipment["_row"], order_id, "volume_m3", actual_volume,
                 f"vypocitany objem {expected_volume}")

error_count = sum(value for (level, _), value in counts.items() if level == "ERROR")
warning_count = sum(value for (level, _), value in counts.items() if level == "WARNING")
sales_dates = [value for row in sales if (value := day(row["order_date"]))]
purchase_dates = [value for row in purchases if (value := day(row["order_date"]))]

print(f"WORKBOOK={path.name}")
for sheet in HEADERS:
    print(f"SHEET={sheet} DATA_ROWS={len(data[sheet])}")
print()
print(f"PRODUCT_COUNT={len(product_by_id)}")
print(f"SALES_ORDER_COUNT={len(sales_by_order)}")
print(f"PURCHASE_ORDER_COUNT={len(purchases_by_order)}")
print(f"INVENTORY_SNAPSHOT_COUNT={len(snapshot_dates)}")
print(f"EXPEDITION_ORDER_COUNT={len(expedition_by_order)}")
print(f"VEHICLE_COUNT={len(vehicle_by_id)}")
if sales_dates:
    print(f"SALES_DATE_MIN={min(sales_dates)}")
    print(f"SALES_DATE_MAX={max(sales_dates)}")
if purchase_dates:
    print(f"PURCHASE_DATE_MIN={min(purchase_dates)}")
    print(f"PURCHASE_DATE_MAX={max(purchase_dates)}")
print()
print(f"BUSINESS_ERROR_COUNT={error_count}")
print(f"BUSINESS_WARNING_COUNT={warning_count}")
for level in ("ERROR", "WARNING"):
    for (issue_level, code), count in sorted(counts.items()):
        if issue_level != level:
            continue
        print()
        print(f"{level}_CODE={code}")
        print(f"{level}_COUNT={count}")
        for example in examples[(issue_level, code)]:
            print(f"{level}_EXAMPLE={example}")
print()
if error_count:
    print("WORKBOOK_BUSINESS_VALID=NIE")
    raise SystemExit(1)
print("WORKBOOK_BUSINESS_VALID=ANO")
