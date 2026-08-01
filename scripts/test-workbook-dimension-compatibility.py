#!/usr/bin/env python3
"""Regression test for XLSX files with incorrect worksheet dimensions."""

from __future__ import annotations

from datetime import date, datetime
import importlib.util
from pathlib import Path
import re
import subprocess
import sys
import tempfile
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent

HEADERS = {
    "products": "product_id product_name category unit purchase_price sales_price supplier minimum_order_quantity lead_time_days weight_kg volume_m3".split(),
    "sales": "order_id order_date product_id quantity unit_price customer_id customer_name region order_status expedition_date".split(),
    "inventory": "snapshot_date product_id stock_actual stock_reserved stock_available warehouse_location min_stock max_stock".split(),
    "purchases": "purchase_order_id order_date delivery_date supplier product_id ordered_quantity delivered_quantity purchase_price".split(),
    "expedition": "order_id received_at picked_at expedition_date delivery_type vehicle_id region weight_kg volume_m3".split(),
    "vehicles": "vehicle_id capacity_kg capacity_m3 availability cost_per_km driver".split(),
}

TYPES = {
    "products": ["text", "text", "text", "text", "decimal(10,2)", "decimal(10,2)", "text", "integer", "integer", "decimal(10,3)", "decimal(10,5)"],
    "sales": ["text", "date", "text", "integer", "decimal(10,2)", "text", "text", "text", "text", "date"],
    "inventory": ["date", "text", "integer", "integer", "integer", "text", "integer", "integer"],
    "purchases": ["text", "date", "date", "text", "text", "integer", "integer", "decimal(10,2)"],
    "expedition": ["text", "datetime", "datetime", "date", "text", "text", "text", "decimal(12,2)", "decimal(12,3)"],
    "vehicles": ["text", "integer", "decimal(10,2)", "text", "decimal(10,2)", "text"],
}

ROWS = {
    "products": ["P1", "Produkt", "Izolacie", "ks", 10.0, 15.0, "Dodavatel", 1, 5, 2.5, 0.01],
    "sales": ["S1", date(2026, 1, 2), "P1", 2, 15.0, "C1", "Zakaznik", "BA", "vybavená", date(2026, 1, 3)],
    "inventory": [date(2026, 1, 1), "P1", 10, 2, 8, "A-01", 3, 20],
    "purchases": ["PO1", date(2026, 1, 1), date(2026, 1, 5), "Dodavatel", "P1", 5, 0, 10.0],
    "expedition": ["S1", datetime(2026, 1, 2, 8, 0), datetime(2026, 1, 2, 9, 0), date(2026, 1, 3), "vlastná", "V1", "BA", 5.0, 0.02],
    "vehicles": ["V1", 1200, 8.0, "pondelok-piatok", 0.55, "Vodic"],
}


def make_fixture(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    dictionary = workbook.create_sheet("data_dictionary")
    dictionary.append(
        ["sheet_name", "column_name", "data_type", "required", "description"]
    )
    for sheet_name, headers in HEADERS.items():
        for column, data_type in zip(headers, TYPES[sheet_name], strict=True):
            required = "nie" if (sheet_name, column) == ("expedition", "vehicle_id") else "áno"
            dictionary.append([sheet_name, column, data_type, required, column])
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(headers)
        worksheet.append(ROWS[sheet_name])
    workbook.save(path)
    workbook.close()


def corrupt_dimensions(source: Path, target: Path) -> None:
    with ZipFile(source, "r") as reader, ZipFile(
        target, "w", compression=ZIP_DEFLATED
    ) as writer:
        for item in reader.infolist():
            payload = reader.read(item.filename)
            if re.fullmatch(r"xl/worksheets/sheet[2-7]\.xml", item.filename):
                text = payload.decode("utf-8")
                text, count = re.subn(
                    r'<dimension ref="[^"]+"\s*/>',
                    "",
                    text,
                    count=1,
                )
                if count != 1:
                    raise RuntimeError(
                        f"dimension element missing in {item.filename}"
                    )
                payload = text.encode("utf-8")
            writer.writestr(item, payload)


def run_validator(script: str, workbook: Path) -> None:
    result = subprocess.run(
        [sys.executable, str(SCRIPT_DIR / script), str(workbook)],
        check=False,
        text=True,
        capture_output=True,
    )
    if result.returncode != 0:
        raise AssertionError(
            f"{script} failed with {result.returncode}\n"
            f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        )


with tempfile.TemporaryDirectory() as temp_directory:
    root = Path(temp_directory)
    valid = root / "valid.xlsx"
    malformed = root / "malformed-dimensions.xlsx"
    make_fixture(valid)
    corrupt_dimensions(valid, malformed)

    readonly = load_workbook(malformed, read_only=True, data_only=False)
    normal = load_workbook(malformed, read_only=False, data_only=False)
    try:
        assert readonly["sales"].max_row is None
        assert normal["sales"].max_row == 2
    finally:
        readonly.close()
        normal.close()

    for validator in (
        "validate-workbook-structure.py",
        "validate-workbook-values.py",
        "validate-workbook-business.py",
    ):
        run_validator(validator, malformed)

    module_path = SCRIPT_DIR / "import-workbook-raw.py"
    spec = importlib.util.spec_from_file_location("import_workbook_raw", module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("cannot load import-workbook-raw.py")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    workbook = load_workbook(malformed, read_only=False, data_only=False)
    try:
        headers, rows = module.read_sheet_rows(workbook["sales"])
    finally:
        workbook.close()
    assert headers == HEADERS["sales"]
    assert len(rows) == 1
    assert rows[0][1]["order_id"] == "S1"

print("WORKBOOK_DIMENSION_COMPATIBILITY_TEST=PASS")
