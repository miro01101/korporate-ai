#!/usr/bin/env python3
from decimal import Decimal
from pathlib import Path
import tempfile
from openpyxl import Workbook, load_workbook
from workbook_normalization import NumericNormalizationError, normalize_numeric, normalize_workbook

CASES = (
    ("153,86 EUR", "decimal(10,2)", Decimal("153.86")),
    ("219,00 EUR", "decimal(10,2)", Decimal("219.00")),
    ("0,55 EUR", "decimal(10,2)", Decimal("0.55")),
    ("1 200", "integer", 1200),
    ("1\u00a0200", "integer", 1200),
    ("1\u202f200", "integer", 1200),
    ("1.200", "integer", 1200),
    ("1.234,56 EUR", "decimal(10,2)", Decimal("1234.56")),
    ("1,234.56 EUR", "decimal(10,2)", Decimal("1234.56")),
)
for value, declared_type, expected in CASES:
    assert normalize_numeric(value, declared_type) == expected

for value, declared_type in (
    ("abc", "decimal(10,2)"),
    ("12 USD", "decimal(10,2)"),
    ("1,2,3", "decimal(10,2)"),
    ("1200,5", "integer"),
    ("NaN", "decimal(10,2)"),
):
    try:
        normalize_numeric(value, declared_type)
    except NumericNormalizationError:
        pass
    else:
        raise AssertionError(value)

with tempfile.TemporaryDirectory() as directory:
    path = Path(directory) / "localized.xlsx"
    workbook = Workbook()
    dictionary = workbook.active
    dictionary.title = "data_dictionary"
    dictionary.append(["sheet_name", "column_name", "data_type", "required", "description"])
    dictionary.append(["products", "product_id", "text", "áno", "id"])
    dictionary.append(["products", "purchase_price", "decimal(10,2)", "áno", "price"])
    dictionary.append(["vehicles", "vehicle_id", "text", "áno", "id"])
    dictionary.append(["vehicles", "capacity_kg", "integer", "áno", "capacity"])
    products = workbook.create_sheet("products")
    products.append(["product_id", "purchase_price"])
    products.append(["P1", "153,86 EUR"])
    vehicles = workbook.create_sheet("vehicles")
    vehicles.append(["vehicle_id", "capacity_kg"])
    vehicles.append(["V1", "1\u00a0200"])
    workbook.save(path)
    workbook.close()
    workbook = load_workbook(path, read_only=False, data_only=False)
    result = normalize_workbook(workbook)
    assert result["normalized_values"] == 2
    assert workbook["products"]["B2"].value == Decimal("153.86")
    assert workbook["vehicles"]["B2"].value == 1200
    workbook.close()
print("WORKBOOK_LOCALIZED_NUMBER_TEST=PASS")
