from pathlib import Path
import sys

from openpyxl import load_workbook


if len(sys.argv) != 2:
    raise SystemExit("Pouzitie: inspect-workbook.py /cesta/subor.xlsx")

path = Path(sys.argv[1])

if not path.is_file():
    raise SystemExit(f"Subor neexistuje: {path}")

workbook = load_workbook(
    filename=path,
    read_only=True,
    data_only=False,
)

print(f"WORKBOOK={path.name}")
print(f"SHEET_COUNT={len(workbook.sheetnames)}")
print(f"SHEET_NAMES={workbook.sheetnames}")

for worksheet in workbook.worksheets:
    print()
    print(f"SHEET={worksheet.title}")
    print(f"STATE={worksheet.sheet_state}")
    print(f"DIMENSION={worksheet.calculate_dimension()}")
    print(f"MAX_ROW={worksheet.max_row}")
    print(f"MAX_COLUMN={worksheet.max_column}")

    max_preview_row = min(worksheet.max_row or 0, 3)
    max_preview_column = min(worksheet.max_column or 0, 20)

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=max_preview_row,
            max_col=max_preview_column,
            values_only=True,
        ),
        start=1,
    ):
        print(f"ROW_{row_number}={list(row)}")

workbook.close()
print()
print("WORKBOOK_INSPECTION_OK=ANO")
