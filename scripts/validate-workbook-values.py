from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import math
from pathlib import Path
import re
import sys

from openpyxl import load_workbook
from workbook_normalization import normalize_workbook


DECIMAL_PATTERN = re.compile(
    r"^decimal\((?P<precision>\d+),(?P<scale>\d+)\)$"
)

REQUIRED_VALUES = {
    "áno",
    "ano",
    "yes",
    "true",
    "1",
}

MAX_EXAMPLES_PER_CODE = 10


if len(sys.argv) != 2:
    raise SystemExit(
        "Pouzitie: validate-workbook-values.py /cesta/subor.xlsx"
    )

path = Path(sys.argv[1])

if not path.is_file():
    raise SystemExit(f"Subor neexistuje: {path}")


issue_counts: Counter[tuple[str, str]] = Counter()
issue_examples: dict[tuple[str, str], list[str]] = defaultdict(list)


def add_issue(
    level: str,
    code: str,
    sheet: str,
    row_number: int,
    column: str,
    message: str,
) -> None:
    key = (level, code)
    issue_counts[key] += 1

    if len(issue_examples[key]) < MAX_EXAMPLES_PER_CODE:
        issue_examples[key].append(
            f"{sheet}!{column}{row_number}: {message}"
        )


def is_blank(value: object) -> bool:
    return value is None or (
        isinstance(value, str) and not value.strip()
    )


def is_integer_value(value: object) -> bool:
    if isinstance(value, bool):
        return False

    if isinstance(value, int):
        return True

    if isinstance(value, float):
        return math.isfinite(value) and value.is_integer()

    if isinstance(value, Decimal):
        return value.is_finite() and value == value.to_integral_value()

    return False


def to_decimal(value: object) -> Decimal | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, Decimal):
        number = value
    elif isinstance(value, int):
        number = Decimal(value)
    elif isinstance(value, float):
        if not math.isfinite(value):
            return None
        number = Decimal(str(value))
    else:
        return None

    if not number.is_finite():
        return None

    return number


def decimal_places(number: Decimal) -> int:
    normalized = number.normalize()

    if normalized == 0:
        return 0

    exponent = normalized.as_tuple().exponent
    return max(-exponent, 0)


def validate_value(
    *,
    sheet: str,
    row_number: int,
    column: str,
    declared_type: str,
    required: bool,
    value: object,
) -> None:
    if is_blank(value):
        if required:
            add_issue(
                "ERROR",
                "REQUIRED_VALUE_MISSING",
                sheet,
                row_number,
                column,
                "povinna hodnota chyba",
            )
        return

    if isinstance(value, str):
        if value.startswith("="):
            add_issue(
                "ERROR",
                "FORMULA_NOT_ALLOWED",
                sheet,
                row_number,
                column,
                "data obsahujú Excel formulu namiesto hodnoty",
            )

        if value != value.strip():
            add_issue(
                "WARNING",
                "TEXT_SURROUNDING_WHITESPACE",
                sheet,
                row_number,
                column,
                "text obsahuje medzeru na zaciatku alebo konci",
            )

    if declared_type == "text":
        if not isinstance(value, str):
            add_issue(
                "ERROR",
                "INVALID_TEXT",
                sheet,
                row_number,
                column,
                f"ocakavany text, ziskany typ {type(value).__name__}",
            )
        return

    if declared_type == "integer":
        if not is_integer_value(value):
            add_issue(
                "ERROR",
                "INVALID_INTEGER",
                sheet,
                row_number,
                column,
                f"ocakavane cele cislo, ziskany typ {type(value).__name__}",
            )
        return

    decimal_match = DECIMAL_PATTERN.fullmatch(declared_type)

    if decimal_match:
        number = to_decimal(value)

        if number is None:
            add_issue(
                "ERROR",
                "INVALID_DECIMAL",
                sheet,
                row_number,
                column,
                f"ocakavane desatinne cislo, ziskany typ {type(value).__name__}",
            )
            return

        precision = int(decimal_match.group("precision"))
        scale = int(decimal_match.group("scale"))

        if decimal_places(number) > scale:
            add_issue(
                "ERROR",
                "DECIMAL_SCALE_EXCEEDED",
                sheet,
                row_number,
                column,
                f"hodnota ma viac ako {scale} desatinnych miest",
            )

        maximum_absolute_value = Decimal(10) ** (precision - scale)

        if abs(number) >= maximum_absolute_value:
            add_issue(
                "ERROR",
                "DECIMAL_PRECISION_EXCEEDED",
                sheet,
                row_number,
                column,
                f"hodnota sa nezmesti do decimal({precision},{scale})",
            )
        return

    if declared_type == "date":
        if not isinstance(value, date):
            add_issue(
                "ERROR",
                "INVALID_DATE",
                sheet,
                row_number,
                column,
                f"ocakavany datum, ziskany typ {type(value).__name__}",
            )
        return

    if declared_type == "datetime":
        if not isinstance(value, datetime):
            add_issue(
                "ERROR",
                "INVALID_DATETIME",
                sheet,
                row_number,
                column,
                f"ocakavany datetime, ziskany typ {type(value).__name__}",
            )
        return

    add_issue(
        "ERROR",
        "UNKNOWN_CONTRACT_TYPE",
        sheet,
        row_number,
        column,
        f"neznamy deklarovany typ {declared_type}",
    )


workbook = load_workbook(
    filename=path,
    read_only=False,
    data_only=False,
)
normalize_workbook(workbook)

if "data_dictionary" not in workbook.sheetnames:
    raise SystemExit("Chyba povinny sheet data_dictionary.")


contract_sheet = workbook["data_dictionary"]
contract: dict[str, list[dict[str, object]]] = defaultdict(list)

for row in contract_sheet.iter_rows(
    min_row=2,
    values_only=True,
):
    sheet_name, column_name, data_type, required, description = row

    if not sheet_name or not column_name:
        continue

    contract[str(sheet_name)].append(
        {
            "column_name": str(column_name),
            "data_type": str(data_type).strip().lower(),
            "required": str(required).strip().lower()
            in REQUIRED_VALUES,
            "description": description,
        }
    )


total_rows = 0
total_values_checked = 0

print(f"WORKBOOK={path.name}")
print(f"CONTRACT_SHEET_COUNT={len(contract)}")

for sheet_name, columns in contract.items():
    worksheet = workbook[sheet_name]

    header_row = next(
        worksheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        )
    )

    header = [
        "" if value is None else str(value)
        for value in header_row
    ]

    column_indexes = {
        column_name: index
        for index, column_name in enumerate(header)
    }

    sheet_rows = 0

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        sheet_rows += 1

        for definition in columns:
            column_name = str(definition["column_name"])
            column_index = column_indexes[column_name]
            value = row[column_index]

            validate_value(
                sheet=sheet_name,
                row_number=row_number,
                column=column_name,
                declared_type=str(definition["data_type"]),
                required=bool(definition["required"]),
                value=value,
            )

            total_values_checked += 1

    total_rows += sheet_rows
    print(f"SHEET={sheet_name} DATA_ROWS={sheet_rows}")


workbook.close()

error_count = sum(
    count
    for (level, _), count in issue_counts.items()
    if level == "ERROR"
)

warning_count = sum(
    count
    for (level, _), count in issue_counts.items()
    if level == "WARNING"
)

print()
print(f"TOTAL_DATA_ROWS={total_rows}")
print(f"TOTAL_VALUES_CHECKED={total_values_checked}")
print(f"VALUE_ERROR_COUNT={error_count}")
print(f"VALUE_WARNING_COUNT={warning_count}")

for level in ("ERROR", "WARNING"):
    for (issue_level, code), count in sorted(issue_counts.items()):
        if issue_level != level:
            continue

        print()
        print(f"{level}_CODE={code}")
        print(f"{level}_COUNT={count}")

        for example in issue_examples[(issue_level, code)]:
            print(f"{level}_EXAMPLE={example}")

if error_count:
    print()
    print("WORKBOOK_BASIC_VALUES_VALID=NIE")
    raise SystemExit(1)

print()
print("WORKBOOK_BASIC_VALUES_VALID=ANO")
