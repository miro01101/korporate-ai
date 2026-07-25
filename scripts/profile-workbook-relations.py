from collections import Counter
from datetime import date, datetime
from pathlib import Path
import sys

from openpyxl import load_workbook


if len(sys.argv) != 2:
    raise SystemExit(
        "Pouzitie: profile-workbook-relations.py /cesta/subor.xlsx"
    )

path = Path(sys.argv[1])

if not path.is_file():
    raise SystemExit(f"Subor neexistuje: {path}")


def normalize_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def duplicate_count(counter):
    return sum(count - 1 for count in counter.values() if count > 1)


workbook = load_workbook(
    filename=path,
    read_only=True,
    data_only=False,
)

print(f"WORKBOOK={path.name}")

# PRODUCTS
worksheet = workbook["products"]
rows = worksheet.iter_rows(min_row=2, values_only=True)

product_ids = Counter()
product_categories = Counter()
product_suppliers = Counter()

for row in rows:
    product_ids[row[0]] += 1
    product_categories[row[2]] += 1
    product_suppliers[row[6]] += 1

product_id_set = set(product_ids)

print()
print("SECTION=PRODUCTS")
print(f"PRODUCT_COUNT={len(product_id_set)}")
print(f"PRODUCT_ID_DUPLICATES={duplicate_count(product_ids)}")
print(f"PRODUCT_CATEGORY_COUNT={len(product_categories)}")
print(f"PRODUCT_SUPPLIER_COUNT={len(product_suppliers)}")

# VEHICLES
worksheet = workbook["vehicles"]
rows = worksheet.iter_rows(min_row=2, values_only=True)

vehicle_ids = Counter()

for row in rows:
    vehicle_ids[row[0]] += 1

vehicle_id_set = set(vehicle_ids)

print()
print("SECTION=VEHICLES")
print(f"VEHICLE_COUNT={len(vehicle_id_set)}")
print(f"VEHICLE_ID_DUPLICATES={duplicate_count(vehicle_ids)}")

# SALES
worksheet = workbook["sales"]
rows = worksheet.iter_rows(min_row=2, values_only=True)

sales_order_ids = Counter()
sales_product_missing = 0
sales_customers = set()
sales_regions = Counter()
sales_statuses = Counter()
sales_dates = []

for row in rows:
    order_id = row[0]
    order_date = normalize_date(row[1])
    product_id = row[2]

    sales_order_ids[order_id] += 1
    sales_customers.add(row[5])
    sales_regions[row[7]] += 1
    sales_statuses[row[8]] += 1

    if product_id not in product_id_set:
        sales_product_missing += 1

    if order_date:
        sales_dates.append(order_date)

sales_order_id_set = set(sales_order_ids)

print()
print("SECTION=SALES")
print(f"SALES_LINE_COUNT={sum(sales_order_ids.values())}")
print(f"SALES_ORDER_COUNT={len(sales_order_id_set)}")
print(f"SALES_CUSTOMER_COUNT={len(sales_customers)}")
print(f"SALES_PRODUCT_FK_MISSING={sales_product_missing}")
print(f"SALES_REGION_VALUES={dict(sorted(sales_regions.items()))}")
print(f"SALES_STATUS_VALUES={dict(sorted(sales_statuses.items()))}")
print(f"SALES_DATE_MIN={min(sales_dates)}")
print(f"SALES_DATE_MAX={max(sales_dates)}")

# INVENTORY
worksheet = workbook["inventory"]
rows = worksheet.iter_rows(min_row=2, values_only=True)

inventory_keys = Counter()
inventory_snapshot_counts = Counter()
inventory_product_missing = 0
inventory_equation_errors = 0
inventory_reserved_over_actual = 0
inventory_min_over_max = 0

for row in rows:
    snapshot_date = normalize_date(row[0])
    product_id = row[1]

    inventory_keys[(snapshot_date, product_id)] += 1
    inventory_snapshot_counts[snapshot_date] += 1

    if product_id not in product_id_set:
        inventory_product_missing += 1

    if row[4] != row[2] - row[3]:
        inventory_equation_errors += 1

    if row[3] > row[2]:
        inventory_reserved_over_actual += 1

    if row[6] > row[7]:
        inventory_min_over_max += 1

print()
print("SECTION=INVENTORY")
print(f"INVENTORY_ROW_COUNT={sum(inventory_snapshot_counts.values())}")
print(f"INVENTORY_KEY_DUPLICATES={duplicate_count(inventory_keys)}")
print(f"INVENTORY_PRODUCT_FK_MISSING={inventory_product_missing}")
print(f"INVENTORY_SNAPSHOT_COUNT={len(inventory_snapshot_counts)}")
print(f"INVENTORY_ROWS_PER_SNAPSHOT_MIN={min(inventory_snapshot_counts.values())}")
print(f"INVENTORY_ROWS_PER_SNAPSHOT_MAX={max(inventory_snapshot_counts.values())}")
print(f"INVENTORY_DATE_MIN={min(inventory_snapshot_counts)}")
print(f"INVENTORY_DATE_MAX={max(inventory_snapshot_counts)}")
print(f"INVENTORY_EQUATION_ERRORS={inventory_equation_errors}")
print(f"INVENTORY_RESERVED_OVER_ACTUAL={inventory_reserved_over_actual}")
print(f"INVENTORY_MIN_OVER_MAX={inventory_min_over_max}")

# PURCHASES
worksheet = workbook["purchases"]
rows = worksheet.iter_rows(min_row=2, values_only=True)

purchase_order_ids = Counter()
purchase_product_missing = 0
purchase_dates = []
purchase_delivery_before_order = 0
purchase_delivered_over_ordered = 0

for row in rows:
    order_date = normalize_date(row[1])
    delivery_date = normalize_date(row[2])
    product_id = row[4]

    purchase_order_ids[row[0]] += 1

    if product_id not in product_id_set:
        purchase_product_missing += 1

    if order_date:
        purchase_dates.append(order_date)

    if order_date and delivery_date and delivery_date < order_date:
        purchase_delivery_before_order += 1

    if row[6] > row[5]:
        purchase_delivered_over_ordered += 1

print()
print("SECTION=PURCHASES")
print(f"PURCHASE_LINE_COUNT={sum(purchase_order_ids.values())}")
print(f"PURCHASE_ORDER_COUNT={len(purchase_order_ids)}")
print(f"PURCHASE_PRODUCT_FK_MISSING={purchase_product_missing}")
print(f"PURCHASE_DATE_MIN={min(purchase_dates)}")
print(f"PURCHASE_DATE_MAX={max(purchase_dates)}")
print(f"PURCHASE_DELIVERY_BEFORE_ORDER={purchase_delivery_before_order}")
print(f"PURCHASE_DELIVERED_OVER_ORDERED={purchase_delivered_over_ordered}")

# EXPEDITION
worksheet = workbook["expedition"]
rows = worksheet.iter_rows(min_row=2, values_only=True)

expedition_order_ids = Counter()
expedition_order_missing = 0
expedition_vehicle_missing = 0
delivery_types = Counter()
expedition_regions = Counter()
picked_before_received = 0

for row in rows:
    order_id = row[0]
    received_at = row[1]
    picked_at = row[2]
    vehicle_id = row[5]

    expedition_order_ids[order_id] += 1
    delivery_types[row[4]] += 1
    expedition_regions[row[6]] += 1

    if order_id not in sales_order_id_set:
        expedition_order_missing += 1

    if vehicle_id is not None and vehicle_id not in vehicle_id_set:
        expedition_vehicle_missing += 1

    if (
        isinstance(received_at, datetime)
        and isinstance(picked_at, datetime)
        and picked_at < received_at
    ):
        picked_before_received += 1

print()
print("SECTION=EXPEDITION")
print(f"EXPEDITION_ROW_COUNT={sum(expedition_order_ids.values())}")
print(f"EXPEDITION_ORDER_DUPLICATES={duplicate_count(expedition_order_ids)}")
print(f"EXPEDITION_ORDER_FK_MISSING={expedition_order_missing}")
print(f"EXPEDITION_VEHICLE_FK_MISSING={expedition_vehicle_missing}")
print(f"DELIVERY_TYPE_VALUES={dict(sorted(delivery_types.items()))}")
print(f"EXPEDITION_REGION_VALUES={dict(sorted(expedition_regions.items()))}")
print(f"PICKED_BEFORE_RECEIVED={picked_before_received}")

workbook.close()

print()
print("RELATIONAL_PROFILE_OK=ANO")
