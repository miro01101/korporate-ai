from collections import defaultdict
from pathlib import Path
import sys

from openpyxl import load_workbook


if len(sys.argv) != 2:
    raise SystemExit(
        "Pouzitie: validate-workbook-structure.py /cesta/subor.xlsx"
    )

path = Path(sys.argv[1])

if not path.is_file():
    raise SystemExit(f"Subor neexistuje: {path}")

workbook = load_workbook(
    filename=path,
    read_only=True,
    data_only=False,
)

errors: list[str] = []

if "data_dictionary" not in workbook.sheetnames:
    raise SystemExit("Chyba povinny sheet data_dictionary.")

dictionary_sheet = workbook["data_dictionary"]
expected_columns: dict[str, list[str]] = defaultdict(list)

for row in dictionary_sheet.iter_rows(
    min_row=2,
    values_only=True,
):
    sheet_name, column_name, *_ = row

    if sheet_name and column_name:
        expected_columns[str(sheet_name)].append(str(column_name))

print(f"WORKBOOK={path.name}")
print(f"CONTRACT_SHEET_COUNT={len(expected_columns)}")

for sheet_name, expected in expected_columns.items():
    print()
    print(f"SHEET={sheet_name}")

    if sheet_name not in workbook.sheetnames:
        errors.append(f"{sheet_name}: sheet chyba")
        print("STATUS=MISSING")
        continue

    worksheet = workbook[sheet_name]

    first_row = next(
        worksheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        )
    )

    actual = [
        str(value) if value is not None else ""
        for value in first_row
    ]

    missing = [
        column for column in expected
        if column not in actual
    ]

    extra = [
        column for column in actual
        if column not in expected
    ]

    duplicate_headers = sorted({
        column for column in actual
        if column and actual.count(column) > 1
    })

    order_matches = actual == expected
    data_rows = max((worksheet.max_row or 1) - 1, 0)

    print(f"DATA_ROWS={data_rows}")
    print(f"EXPECTED_COLUMNS={expected}")
    print(f"ACTUAL_COLUMNS={actual}")
    print(f"MISSING_COLUMNS={missing}")
    print(f"EXTRA_COLUMNS={extra}")
    print(f"DUPLICATE_HEADERS={duplicate_headers}")
    print(f"COLUMN_ORDER_MATCH={order_matches}")

    if missing:
        errors.append(f"{sheet_name}: chybaju stlpce {missing}")

    if extra:
        errors.append(f"{sheet_name}: navyse stlpce {extra}")

    if duplicate_headers:
        errors.append(
            f"{sheet_name}: duplicitne hlavicky {duplicate_headers}"
        )

    if not order_matches:
        errors.append(f"{sheet_name}: nesedi poradie stlpcov")

workbook.close()

print()
print(f"STRUCTURE_ERROR_COUNT={len(errors)}")

for error in errors:
    print(f"ERROR={error}")

if errors:
    print("WORKBOOK_STRUCTURE_VALID=NIE")
    raise SystemExit(1)

print("WORKBOOK_STRUCTURE_VALID=ANO")
