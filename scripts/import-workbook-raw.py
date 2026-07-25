"""Load a validated XLSX workbook into the raw database layer."""

from __future__ import annotations

import argparse
from datetime import date, datetime
from decimal import Decimal
import hashlib
from pathlib import Path
import sys
import uuid

import psycopg
from psycopg.errors import UniqueViolation
from psycopg.types.json import Jsonb
from openpyxl import load_workbook


SHEETS = {
    "products": "xlsx_products",
    "sales": "xlsx_sales",
    "inventory": "xlsx_inventory",
    "purchases": "xlsx_purchases",
    "expedition": "xlsx_expedition",
    "vehicles": "xlsx_vehicles",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load XLSX business sheets into raw PostgreSQL tables."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--db-host", default="postgres")
    parser.add_argument("--db-port", type=int, default=5432)
    parser.add_argument("--db-name", default="korporate_ai")
    parser.add_argument("--db-user", default="korporate_app")
    parser.add_argument(
        "--db-password-file",
        type=Path,
        required=True,
    )
    parser.add_argument(
        "--created-by",
        default="manual-cli",
    )
    parser.add_argument(
        "--source-path",
        default=None,
        help="Host-visible source path stored in audit metadata.",
    )
    parser.add_argument(
        "--expected-sha256",
        default=None,
    )
    return parser.parse_args()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def read_sheet_rows(
    worksheet,
) -> tuple[list[str], list[tuple[int, dict[str, object]]]]:
    header_values = next(
        worksheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        )
    )
    headers = [
        "" if value is None else str(value)
        for value in header_values
    ]

    if not headers or any(not header for header in headers):
        raise ValueError(
            f"Sheet {worksheet.title!r} ma prazdnu hlavicku."
        )

    if len(headers) != len(set(headers)):
        raise ValueError(
            f"Sheet {worksheet.title!r} ma duplicitne hlavicky."
        )

    rows: list[tuple[int, dict[str, object]]] = []

    for row_number, values in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        if all(value is None for value in values):
            continue

        payload = {
            header: json_value(value)
            for header, value in zip(headers, values, strict=True)
        }
        rows.append((row_number, payload))

    return headers, rows


def existing_batch(cursor, file_sha256: str):
    cursor.execute(
        """
        SELECT id, status, row_count_raw
        FROM audit.import_batches
        WHERE file_sha256 = %s
        """,
        (file_sha256,),
    )
    return cursor.fetchone()


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.resolve()

    if not workbook_path.is_file():
        print(f"CHYBA: workbook neexistuje: {workbook_path}", file=sys.stderr)
        return 2

    if not args.db_password_file.is_file():
        print(
            "CHYBA: subor s databazovym heslom neexistuje: "
            f"{args.db_password_file}",
            file=sys.stderr,
        )
        return 2

    file_sha256 = sha256_file(workbook_path)
    if (
        args.expected_sha256
        and file_sha256 != args.expected_sha256.lower()
    ):
        print(
            "CHYBA: SHA-256 workbooku sa nezhoduje s ocakavanim.",
            file=sys.stderr,
        )
        print(f"ACTUAL_SHA256={file_sha256}", file=sys.stderr)
        return 2

    workbook = load_workbook(
        filename=workbook_path,
        read_only=True,
        data_only=False,
    )

    missing_sheets = sorted(set(SHEETS) - set(workbook.sheetnames))
    if missing_sheets:
        workbook.close()
        print(
            "CHYBA: chybaju sheety: " + ", ".join(missing_sheets),
            file=sys.stderr,
        )
        return 2

    sheet_rows: dict[str, list[tuple[int, dict[str, object]]]] = {}
    sheet_headers: dict[str, list[str]] = {}

    try:
        for sheet_name in SHEETS:
            headers, rows = read_sheet_rows(workbook[sheet_name])
            sheet_headers[sheet_name] = headers
            sheet_rows[sheet_name] = rows
    finally:
        workbook.close()

    total_rows = sum(len(rows) for rows in sheet_rows.values())
    password = args.db_password_file.read_text(encoding="utf-8").strip()
    source_path = args.source_path or str(workbook_path)
    batch_id = uuid.uuid4()

    connection = psycopg.connect(
        host=args.db_host,
        port=args.db_port,
        dbname=args.db_name,
        user=args.db_user,
        password=password,
    )

    try:
        with connection.cursor() as cursor:
            duplicate = existing_batch(cursor, file_sha256)
            if duplicate:
                connection.rollback()
                print("BATCH_DUPLICATE=ANO")
                print(f"IMPORT_BATCH_ID={duplicate[0]}")
                print(f"IMPORT_BATCH_STATUS={duplicate[1]}")
                print(f"RAW_ROW_COUNT={duplicate[2]}")
                return 0

            cursor.execute(
                """
                SELECT
                    max(value) FILTER (WHERE key = 'platform_version'),
                    max(value) FILTER (WHERE key = 'schema_revision')
                FROM meta.system_info
                WHERE key IN ('platform_version', 'schema_revision')
                """
            )
            platform_version, schema_revision = cursor.fetchone()

            try:
                cursor.execute(
                    """
                    INSERT INTO audit.import_batches (
                        id,
                        source_type,
                        original_filename,
                        source_path,
                        file_sha256,
                        file_size_bytes,
                        platform_version,
                        schema_revision,
                        status,
                        created_by,
                        metadata
                    )
                    VALUES (
                        %s,
                        'manual_xlsx',
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        'registered',
                        %s,
                        %s
                    )
                    """,
                    (
                        batch_id,
                        workbook_path.name,
                        source_path,
                        file_sha256,
                        workbook_path.stat().st_size,
                        platform_version,
                        schema_revision,
                        args.created_by,
                        Jsonb(
                            {
                                "loader": "scripts/import-workbook-raw.py",
                                "sheet_headers": sheet_headers,
                            }
                        ),
                    ),
                )
                connection.commit()
            except UniqueViolation:
                connection.rollback()
                duplicate = existing_batch(cursor, file_sha256)
                connection.rollback()
                print("BATCH_DUPLICATE=ANO")
                print(f"IMPORT_BATCH_ID={duplicate[0]}")
                print(f"IMPORT_BATCH_STATUS={duplicate[1]}")
                print(f"RAW_ROW_COUNT={duplicate[2]}")
                return 0

            cursor.execute(
                """
                UPDATE audit.import_batches
                SET status = 'validating'
                WHERE id = %s
                """,
                (batch_id,),
            )
            connection.commit()

            try:
                for sheet_name, table_name in SHEETS.items():
                    records = [
                        (
                            batch_id,
                            row_number,
                            Jsonb(payload),
                        )
                        for row_number, payload in sheet_rows[sheet_name]
                    ]
                    cursor.executemany(
                        f"""
                        INSERT INTO raw.{table_name} (
                            import_batch_id,
                            source_row_number,
                            source_data
                        )
                        VALUES (%s, %s, %s)
                        """,
                        records,
                    )

                cursor.execute(
                    """
                    UPDATE audit.import_batches
                    SET status = 'raw_loaded',
                        row_count_raw = %s,
                        metadata = metadata || %s
                    WHERE id = %s
                    """,
                    (
                        total_rows,
                        Jsonb(
                            {
                                "sheet_row_counts": {
                                    name: len(rows)
                                    for name, rows in sheet_rows.items()
                                }
                            }
                        ),
                        batch_id,
                    ),
                )
                connection.commit()
            except Exception as exc:
                connection.rollback()
                cursor.execute(
                    """
                    UPDATE audit.import_batches
                    SET status = 'failed',
                        finished_at = now(),
                        error_message = %s
                    WHERE id = %s
                    """,
                    (str(exc)[:4000], batch_id),
                )
                connection.commit()
                raise

    finally:
        connection.close()

    print("BATCH_DUPLICATE=NIE")
    print(f"IMPORT_BATCH_ID={batch_id}")
    print(f"FILE_SHA256={file_sha256}")
    for sheet_name in SHEETS:
        print(
            f"SHEET={sheet_name} "
            f"RAW_ROWS={len(sheet_rows[sheet_name])}"
        )
    print(f"RAW_ROW_COUNT={total_rows}")
    print("IMPORT_BATCH_STATUS=raw_loaded")
    print("RAW_IMPORT_OK=ANO")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
